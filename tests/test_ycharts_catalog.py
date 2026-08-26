"""Tests for the local YCharts Excel reference parser and search."""

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from services.ycharts_catalog import YChartsCatalog, YChartsCatalogError


def make_reference(path: Path) -> None:
    workbook = Workbook()
    metrics = workbook.active
    metrics.title = "Company | Metrics"
    metrics.append(["YCharts Excel Add-In Company Metrics"])
    metrics.append([])
    metrics.append(["Metric Name", "Financial Statement", "Metric Code", "Syntax"])
    metrics.append(["Year to Date Total Returns (Daily)", "", "ytd_total_return", 'YCP("MSFT","ytd_total_return")'])
    metrics.append(["Dividend Yield", "", "dividend_yield", 'YCP("MSFT","dividend_yield")'])
    info = workbook.create_sheet("Company | Info")
    info.append(["YCharts Excel Add-In Company Info"])
    info.append([])
    info.append(["Metric Name", "Metric Code", "Type", "Description", "Syntax"])
    info.append(["Security Name", "security_name", "String", "Security name", 'YCI("MSFT","security_name")'])
    workbook.save(path)


class YChartsCatalogTests(unittest.TestCase):
    def test_reads_metric_codes_from_local_reference(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "reference.xlsx"
            make_reference(path)
            catalog = YChartsCatalog.from_reference_workbook(path)

            self.assertEqual(len(catalog.metrics), 3)
            self.assertTrue(catalog.contains("YCP", "ytd_total_return"))
            self.assertTrue(catalog.contains("YCS", "ytd_total_return"))
            self.assertTrue(catalog.contains("YCI", "security_name"))
            self.assertFalse(catalog.contains("YCP", "invented_metric"))

    def test_search_finds_plain_language_performance_metrics(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "reference.xlsx"
            make_reference(path)
            matches = YChartsCatalog.from_reference_workbook(path).search(
                "Compare YTD performance and dividend yield", limit=20
            )
            codes = {item.code for item in matches}

            self.assertIn("ytd_total_return", codes)
            self.assertIn("dividend_yield", codes)

    def test_rejects_unrelated_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "unrelated.xlsx"
            workbook = Workbook()
            workbook.active.append(["Not a YCharts reference"])
            workbook.save(path)

            with self.assertRaisesRegex(YChartsCatalogError, "not the YCharts"):
                YChartsCatalog.from_reference_workbook(path)


if __name__ == "__main__":
    unittest.main()
