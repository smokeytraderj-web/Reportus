import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from pypdf import PdfReader
from reportlab.pdfgen import canvas

from generators.portfolio_pdf import (
    build_portfolio_workbook_pdf,
    is_portfolio_workbook,
    load_portfolio_workbook,
)


def _make_portfolio(path: Path) -> None:
    workbook = Workbook()
    portfolio = workbook.active
    portfolio.title = "Portfolio"
    portfolio["A1"] = "Test Long-Term Portfolio"
    portfolio["A2"] = "Benchmark: VT | As of 2026-08-26 | USD brokerage account"
    portfolio["A4"] = "Thesis"
    portfolio["A5"] = "Combine diversified factor exposure with selected compounders."
    portfolio.append([])
    portfolio.append([])
    portfolio.append([
        "Ticker", "Holding", "Sleeve", "Price", "Target %", "Invest now",
        "Monthly", "Expense ratio", "Valuation", "One-line reason",
    ])
    portfolio.append(["AAA", "Example ETF", "Factor ETF", 50, .60, 60, 60, .003, "Factor portfolio", "Diversified exposure."])
    portfolio.append(["BBB", "Example Company", "Selected compounder", 100, .40, 40, 40, 0, "20x P/E", "Durable reinvestment runway."])
    portfolio.append(["TOTAL", None, None, None, 1, 100, 100, .0018])
    portfolio.append([])
    portfolio.append(["Biggest risk"])
    portfolio.append(["Valuation compression across the selected holdings."])
    portfolio.append([])
    portfolio.append(["Implementation rule: invest monthly and use new money to correct drift."])

    assumptions = workbook.create_sheet("Research & Assumptions")
    assumptions.append(["Research, Assumptions & Controls"])
    assumptions.append([])
    assumptions.append(["Initial contribution", 100])
    assumptions.append(["Monthly contribution", 100])
    assumptions.append(["Base currency", "USD"])
    assumptions.append(["Account assumption", "Taxable brokerage with fractional shares"])
    assumptions.append(["Benchmark", "VT"])
    assumptions.append(["Price timestamp", "2026-08-26"])
    assumptions.append([])
    assumptions.append(["Recent-development sanity check"])
    assumptions.append(["Ticker", "Recent material development", "Portfolio interpretation", "Source ID"])
    assumptions.append(["BBB", "Revenue increased.", "Monitor valuation discipline.", "S01"])

    sources = workbook.create_sheet("Sources")
    sources.append(["Sources & Data Audit"])
    sources.append([])
    sources.append(["ID", "Item", "As-of / published", "Source", "URL", "What it supports"])
    sources.append(["S01", "Example Company", "2026-08-26", "Issuer", "https://example.com", "Recent results"])
    workbook.save(path)


def _make_reference(path: Path) -> None:
    document = canvas.Canvas(str(path), pagesize=(960, 540))
    document.drawString(72, 468, "Approved widescreen reference")
    document.save()


class PortfolioPDFTests(unittest.TestCase):
    def test_detects_and_maps_real_portfolio_headers(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "portfolio.xlsx"
            _make_portfolio(source)

            self.assertTrue(is_portfolio_workbook(source))
            data = load_portfolio_workbook(source)

            self.assertEqual(tuple(item.ticker for item in data.holdings), ("AAA", "BBB"))
            self.assertAlmostEqual(sum(item.target for item in data.holdings), 1.0)
            self.assertEqual(data.holdings[0].sleeve, "Factor ETF")
            self.assertEqual(len(data.research_updates), 1)
            self.assertEqual(len(data.sources), 1)

    def test_builds_reference_sized_portfolio_pdf(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "portfolio.xlsx"
            reference = root / "reference.pdf"
            output = root / "report.pdf"
            _make_portfolio(source)
            _make_reference(reference)

            build_portfolio_workbook_pdf(
                source,
                reference,
                output,
                client_name="Sample Household",
                report_title="Portfolio Review",
                period_label="August 2026",
                source_label="Uploaded portfolio workbook",
            )

            reader = PdfReader(output)
            self.assertGreaterEqual(len(reader.pages), 5)
            self.assertAlmostEqual(float(reader.pages[0].mediabox.width), 960)
            self.assertAlmostEqual(float(reader.pages[0].mediabox.height), 540)


if __name__ == "__main__":
    unittest.main()
