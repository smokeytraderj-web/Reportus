"""Structured, branded custom Excel workbooks with validated YCharts formulas."""

from __future__ import annotations

import datetime as dt
import math
import re
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, DoughnutChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from quality.output_qa import OutputInspector
from services.ycharts_catalog import YChartsCatalog


NAVY = "1B2A4A"
NAVY_DARK = "0A1224"
GOLD = "BFA054"
INK = "172434"
MUTED = "667487"
LIGHT = "F3F5F8"
RULE = "D8DEE6"
WHITE = "FFFFFF"
GREEN = "1F7A4D"
RED = "B42318"
FIRM_NAME = "Gottfried & Somberg Wealth Management"

FORMATS = {
    "text": "@",
    "integer": "#,##0;[Red](#,##0);-",
    "decimal": "#,##0.00;[Red](#,##0.00);-",
    "currency": "$#,##0.00;[Red]($#,##0.00);-",
    "percent": "0.00%;[Red](0.00%);-",
    "multiple": "0.0x;[Red](0.0x);-",
    "date": "mmm d, yyyy",
}
YCHARTS_FUNCTIONS = frozenset({"YCP", "YCI", "YCS", "YCD", "YCDS", "YCH", "YCU"})
YCU_CODES = frozenset({
    "account", "client", "group", "holding", "holding_lot", "related_securities",
    "transaction", "watchlist",
})
SAFE_EXCEL_FUNCTIONS = frozenset({
    "ABS", "AND", "AVERAGE", "AVERAGEIF", "AVERAGEIFS", "COUNT", "COUNTA",
    "COUNTIF", "COUNTIFS", "DATE", "DAY", "IF", "IFERROR", "IFS", "INDEX",
    "LEFT", "LEN", "MATCH", "MAX", "MEDIAN", "MID", "MIN", "MONTH", "NOT",
    "OR", "RIGHT", "ROUND", "ROUNDDOWN", "ROUNDUP", "SUM", "SUMIF", "SUMIFS",
    "TEXT", "TRIM", "XLOOKUP", "YEAR",
})
_SHEET_FORBIDDEN = re.compile(r"[\\/*?:\[\]]")
_SECURITY = re.compile(r"[A-Za-z0-9:^._-]{1,40}")
_FUNCTION = re.compile(r"(?<![A-Za-z0-9_.])([A-Za-z_][A-Za-z0-9_.]*)\s*\(")
_CELL_REFERENCE = re.compile(r"\$?[A-Z]{1,3}\$?\d+")


class CustomWorkbookError(RuntimeError):
    pass


def custom_workbook_schema() -> dict[str, object]:
    cell = {
        "type": "object",
        "required": ["kind"],
        "properties": {
            "kind": {"type": "string", "enum": ["value", "excel_formula", "ycharts"]},
            "value": {"type": ["string", "number", "boolean", "null"]},
            "formula": {"type": "string", "maxLength": 500},
            "function": {"type": "string", "enum": sorted(YCHARTS_FUNCTIONS)},
            "security": {"type": "string", "maxLength": 40},
            "metric_code": {"type": "string", "maxLength": 120},
            "metric_name": {"type": "string", "maxLength": 100},
            "as_of_date": {"type": "string", "maxLength": 10},
            "start_date": {"type": "string", "maxLength": 10},
            "end_date": {"type": "string", "maxLength": 10},
            "last_periods": {"type": "integer", "minimum": -1000, "maximum": -1},
        },
    }
    return {
        "type": "object",
        "required": ["workbook_subtitle", "sheets"],
        "properties": {
            "workbook_subtitle": {"type": "string", "maxLength": 160},
            "sheets": {
                "type": "array", "minItems": 1, "maxItems": 6,
                "items": {
                    "type": "object",
                    "required": ["name", "title", "subtitle", "source_note", "columns", "rows", "charts"],
                    "properties": {
                        "name": {"type": "string", "maxLength": 31},
                        "title": {"type": "string", "maxLength": 90},
                        "subtitle": {"type": "string", "maxLength": 160},
                        "source_note": {"type": "string", "maxLength": 180},
                        "columns": {
                            "type": "array", "minItems": 2, "maxItems": 12,
                            "items": {
                                "type": "object", "required": ["header", "format", "width"],
                                "properties": {
                                    "header": {"type": "string", "maxLength": 55},
                                    "format": {"type": "string", "enum": sorted(FORMATS)},
                                    "width": {"type": "integer", "minimum": 10, "maximum": 42},
                                },
                            },
                        },
                        "rows": {
                            "type": "array", "minItems": 1, "maxItems": 200,
                            "items": {
                                "type": "object", "required": ["cells"],
                                "properties": {"cells": {"type": "array", "minItems": 2, "maxItems": 12, "items": cell}},
                            },
                        },
                        "charts": {
                            "type": "array", "maxItems": 4,
                            "items": {
                                "type": "object",
                                "required": ["type", "title", "category_column", "series_columns", "max_rows"],
                                "properties": {
                                    "type": {"type": "string", "enum": ["area", "bar", "column", "doughnut", "line", "pie"]},
                                    "title": {"type": "string", "maxLength": 90},
                                    "category_column": {"type": "integer", "minimum": 0, "maximum": 11},
                                    "series_columns": {"type": "array", "minItems": 1, "maxItems": 6, "items": {"type": "integer", "minimum": 0, "maximum": 11}},
                                    "max_rows": {"type": "integer", "minimum": 5, "maximum": 500},
                                },
                            },
                        },
                    },
                },
            },
        },
    }


def _text(value: object, field: str, limit: int, *, allow_empty: bool = False) -> str:
    if not isinstance(value, str):
        raise CustomWorkbookError(f"{field} must be text.")
    result = value.strip()
    if not result and not allow_empty:
        raise CustomWorkbookError(f"{field} is required.")
    if len(result) > limit:
        raise CustomWorkbookError(f"{field} is too long.")
    return result


def _date(value: object, field: str, *, required: bool = False) -> str:
    if value in (None, "") and not required:
        return ""
    text = _text(value, field, 10)
    try:
        dt.date.fromisoformat(text)
    except ValueError as exc:
        raise CustomWorkbookError(f"{field} must use YYYY-MM-DD.") from exc
    return text


def _safe_formula(value: object, field: str) -> str:
    formula = _text(value, field, 500)
    if not formula.startswith("="):
        raise CustomWorkbookError(f"{field} must begin with =.")
    if any(token in formula for token in ("[", "]", "|", "{", "}", "\n", "\r")):
        raise CustomWorkbookError(f"{field} contains an unsafe external or array reference.")
    functions = {match.group(1).upper() for match in _FUNCTION.finditer(formula)}
    unsupported = sorted(functions - SAFE_EXCEL_FUNCTIONS)
    if unsupported:
        raise CustomWorkbookError(f"{field} uses unsupported function(s): {', '.join(unsupported)}.")
    return formula


def _normalize_ycharts(cell: dict[str, object], field: str, catalog: YChartsCatalog) -> dict[str, object]:
    function = _text(cell.get("function"), f"{field} function", 5).upper()
    if function not in YCHARTS_FUNCTIONS:
        raise CustomWorkbookError(f"{field} uses an unsupported YCharts function.")
    code = _text(cell.get("metric_code", ""), f"{field} metric code", 120, allow_empty=True)
    security = _text(cell.get("security", ""), f"{field} security", 40, allow_empty=True)
    if security and _SECURITY.fullmatch(security) is None:
        raise CustomWorkbookError(f"{field} security contains unsupported characters.")
    if function == "YCU":
        if code.casefold() not in YCU_CODES:
            raise CustomWorkbookError(f"{field} uses an unknown YCU metric code.")
        security = ""
    elif function == "YCH":
        if not security:
            raise CustomWorkbookError(f"{field} requires a security or portfolio identifier.")
        code = ""
    else:
        macro_point = function == "YCP" and security == "I:USGDP" and not code
        if not security or (not code and not macro_point):
            raise CustomWorkbookError(f"{field} requires a security and metric code.")
        if code and not catalog.contains(function, code):
            raise CustomWorkbookError(f"{field} metric code is not in the supplied YCharts reference.")

    start_date = _date(cell.get("start_date"), f"{field} start date")
    end_date = _date(cell.get("end_date"), f"{field} end date")
    as_of_date = _date(cell.get("as_of_date"), f"{field} as-of date")
    last_periods = cell.get("last_periods")
    if last_periods is not None and (isinstance(last_periods, bool) or not isinstance(last_periods, int) or not -1000 <= last_periods <= -1):
        raise CustomWorkbookError(f"{field} last_periods must be an integer from -1000 to -1.")
    if function in {"YCS", "YCDS"} and not start_date and last_periods is None:
        raise CustomWorkbookError(f"{field} requires start_date or last_periods for a historical series.")
    if end_date and not start_date:
        raise CustomWorkbookError(f"{field} cannot use end_date without start_date.")
    if start_date and end_date and end_date < start_date:
        raise CustomWorkbookError(f"{field} end_date precedes start_date.")
    return {
        "kind": "ycharts", "function": function, "security": security, "metric_code": code,
        "metric_name": _text(cell.get("metric_name", ""), f"{field} metric name", 100, allow_empty=True),
        "as_of_date": as_of_date, "start_date": start_date, "end_date": end_date,
        "last_periods": last_periods,
    }


def normalize_custom_workbook_payload(payload: dict[str, object], *, report_title: str,
                                      catalog: YChartsCatalog) -> dict[str, object]:
    if not isinstance(payload, dict):
        raise CustomWorkbookError("The custom workbook plan must be one JSON object.")
    title = _text(report_title, "Workbook title", 90)
    subtitle = _text(payload.get("workbook_subtitle"), "Workbook subtitle", 160)
    raw_sheets = payload.get("sheets")
    if not isinstance(raw_sheets, list) or not 1 <= len(raw_sheets) <= 6:
        raise CustomWorkbookError("A custom workbook requires 1-6 report sheets.")
    sheets: list[dict[str, object]] = []
    names: set[str] = {"cover"}
    for sheet_number, raw_sheet in enumerate(raw_sheets, start=1):
        if not isinstance(raw_sheet, dict):
            raise CustomWorkbookError(f"Sheet {sheet_number} must be an object.")
        name = _text(raw_sheet.get("name"), f"Sheet {sheet_number} name", 31)
        if _SHEET_FORBIDDEN.search(name) or name.casefold() in names:
            raise CustomWorkbookError(f"Sheet {sheet_number} name is invalid or duplicated.")
        names.add(name.casefold())
        raw_columns = raw_sheet.get("columns")
        if not isinstance(raw_columns, list) or not 2 <= len(raw_columns) <= 12:
            raise CustomWorkbookError(f"Sheet {sheet_number} requires 2-12 columns.")
        columns = []
        for column_number, raw_column in enumerate(raw_columns, start=1):
            if not isinstance(raw_column, dict):
                raise CustomWorkbookError(f"Sheet {sheet_number} column {column_number} must be an object.")
            cell_format = str(raw_column.get("format", "")).casefold()
            width = raw_column.get("width")
            if cell_format not in FORMATS:
                raise CustomWorkbookError(f"Sheet {sheet_number} column {column_number} has an invalid format.")
            if isinstance(width, bool) or not isinstance(width, int) or not 10 <= width <= 42:
                raise CustomWorkbookError(f"Sheet {sheet_number} column {column_number} width is invalid.")
            columns.append({
                "header": _text(raw_column.get("header"), f"Sheet {sheet_number} column {column_number} header", 55),
                "format": cell_format, "width": width,
            })
        raw_rows = raw_sheet.get("rows")
        if not isinstance(raw_rows, list) or not 1 <= len(raw_rows) <= 200:
            raise CustomWorkbookError(f"Sheet {sheet_number} requires 1-200 rows.")
        rows = []
        for row_number, raw_row in enumerate(raw_rows, start=1):
            raw_cells = raw_row.get("cells") if isinstance(raw_row, dict) else None
            if not isinstance(raw_cells, list) or len(raw_cells) != len(columns):
                raise CustomWorkbookError(f"Sheet {sheet_number} row {row_number} has the wrong cell count.")
            cells = []
            for column_number, raw_cell in enumerate(raw_cells, start=1):
                field = f"Sheet {sheet_number} row {row_number} cell {column_number}"
                if not isinstance(raw_cell, dict):
                    raise CustomWorkbookError(f"{field} must be an object.")
                kind = str(raw_cell.get("kind", "")).casefold()
                if kind == "value":
                    value = raw_cell.get("value")
                    if isinstance(value, (dict, list)):
                        raise CustomWorkbookError(f"{field} value must be a scalar.")
                    if isinstance(value, float) and not math.isfinite(value):
                        raise CustomWorkbookError(f"{field} value must be finite.")
                    if isinstance(value, str) and len(value) > 500:
                        raise CustomWorkbookError(f"{field} value is too long.")
                    cells.append({"kind": "value", "value": value})
                elif kind == "excel_formula":
                    cells.append({"kind": "excel_formula", "formula": _safe_formula(raw_cell.get("formula"), field)})
                elif kind == "ycharts":
                    cells.append(_normalize_ycharts(raw_cell, field, catalog))
                else:
                    raise CustomWorkbookError(f"{field} has an unsupported kind.")
            spill_cells = [
                cell for cell in cells
                if cell["kind"] == "ycharts"
                and cell["function"] in {"YCS", "YCDS", "YCH", "YCU"}
            ]
            if spill_cells and row_number != len(raw_rows):
                raise CustomWorkbookError(
                    f"Sheet {sheet_number} historical or list formulas must be in the final row."
                )
            if any(cell["function"] in {"YCH", "YCU"} for cell in spill_cells):
                occupied = [
                    cell for cell in cells
                    if cell not in spill_cells
                    and not (cell["kind"] == "value" and cell.get("value") in (None, ""))
                ]
                if len(spill_cells) != 1 or occupied:
                    raise CustomWorkbookError(
                        f"Sheet {sheet_number} YCH/YCU output requires its own final row."
                    )
            rows.append({"cells": cells})

        raw_charts = raw_sheet.get("charts")
        if not isinstance(raw_charts, list) or len(raw_charts) > 4:
            raise CustomWorkbookError(f"Sheet {sheet_number} charts must be a list with at most four items.")
        charts = []
        for chart_number, raw_chart in enumerate(raw_charts, start=1):
            if not isinstance(raw_chart, dict):
                raise CustomWorkbookError(f"Sheet {sheet_number} chart {chart_number} must be an object.")
            chart_type = str(raw_chart.get("type", "")).casefold()
            category = raw_chart.get("category_column")
            series = raw_chart.get("series_columns")
            max_rows = raw_chart.get("max_rows")
            if chart_type not in {"area", "bar", "column", "doughnut", "line", "pie"}:
                raise CustomWorkbookError(f"Sheet {sheet_number} chart {chart_number} type is invalid.")
            if isinstance(category, bool) or not isinstance(category, int) or not 0 <= category < len(columns):
                raise CustomWorkbookError(f"Sheet {sheet_number} chart {chart_number} category column is invalid.")
            if not isinstance(series, list) or not 1 <= len(series) <= 6 or any(
                isinstance(value, bool) or not isinstance(value, int) or not 0 <= value < len(columns)
                for value in series
            ):
                raise CustomWorkbookError(f"Sheet {sheet_number} chart {chart_number} series columns are invalid.")
            if category in series or len(series) != len(set(series)):
                raise CustomWorkbookError(f"Sheet {sheet_number} chart {chart_number} columns overlap.")
            if any(columns[value]["format"] in {"text", "date"} for value in series):
                raise CustomWorkbookError(
                    f"Sheet {sheet_number} chart {chart_number} series must use numeric columns."
                )
            if isinstance(max_rows, bool) or not isinstance(max_rows, int) or not 5 <= max_rows <= 500:
                raise CustomWorkbookError(f"Sheet {sheet_number} chart {chart_number} max_rows is invalid.")
            charts.append({
                "type": chart_type,
                "title": _text(raw_chart.get("title"), f"Sheet {sheet_number} chart {chart_number} title", 90),
                "category_column": category,
                "series_columns": series,
                "max_rows": max_rows,
            })
        sheets.append({
            "name": name,
            "title": _text(raw_sheet.get("title"), f"Sheet {sheet_number} title", 90),
            "subtitle": _text(raw_sheet.get("subtitle"), f"Sheet {sheet_number} subtitle", 160),
            "source_note": _text(raw_sheet.get("source_note"), f"Sheet {sheet_number} source note", 180),
            "columns": columns, "rows": rows, "charts": charts,
        })
    return {"title": title, "workbook_subtitle": subtitle, "sheets": sheets}


def _quoted(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def ycharts_formula(cell: dict[str, object]) -> str:
    function = str(cell["function"])
    code = str(cell.get("metric_code", ""))
    security = str(cell.get("security", ""))
    if function == "YCU":
        return f"=YCU({_quoted(code)})"
    if function == "YCH":
        return f"=YCH({_quoted(security)})"
    args = [_quoted(security)]
    if code:
        args.append(_quoted(code))
    as_of = str(cell.get("as_of_date", ""))
    start = str(cell.get("start_date", ""))
    end = str(cell.get("end_date", ""))
    last_periods = cell.get("last_periods")
    if function in {"YCS", "YCDS"}:
        if last_periods is not None:
            args.append(str(last_periods))
        else:
            args.append(_quoted(start))
            if end:
                args.append(_quoted(end))
    elif as_of:
        args.append(_quoted(as_of))
    return f"={function}({','.join(args)})"


def _display_value(cell: dict[str, object]) -> str:
    if cell["kind"] == "value":
        value = cell.get("value")
        return "" if value is None else str(value)
    if cell["kind"] == "excel_formula":
        return str(cell["formula"])
    name = str(cell.get("metric_name", "")).strip()
    return name or ycharts_formula(cell).lstrip("=")


def _style_sheet_header(sheet, title: str, subtitle: str, source_note: str, last_column: int) -> None:
    end = get_column_letter(last_column)
    sheet.merge_cells(f"A1:{end}1")
    sheet.merge_cells(f"A2:{end}2")
    sheet.merge_cells(f"A3:{end}3")
    sheet["A1"] = title
    sheet["A1"].font = Font(name="Georgia", size=18, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY_DARK)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(name="Aptos", size=10, color=WHITE)
    sheet["A2"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 26
    sheet["A3"] = f"Source: {source_note}"
    sheet["A3"].font = Font(name="Aptos", size=8, italic=True, color=MUTED)
    sheet["A3"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[3].height = 19
    sheet.row_dimensions[4].height = 8


def _cover(workbook: Workbook, plan: dict[str, object], request_text: str) -> None:
    sheet = workbook.active
    sheet.title = "Cover"
    sheet.sheet_view.showGridLines = False
    for column, width in enumerate((4, 22, 22, 22, 22, 4), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.merge_cells("B2:E3")
    sheet["B2"] = FIRM_NAME
    sheet["B2"].font = Font(name="Georgia", size=20, bold=True, color=WHITE)
    sheet["B2"].fill = PatternFill("solid", fgColor=NAVY_DARK)
    sheet["B2"].alignment = Alignment(vertical="center")
    sheet.merge_cells("B4:E5")
    sheet["B4"] = plan["title"]
    sheet["B4"].font = Font(name="Georgia", size=24, bold=True, color=NAVY)
    sheet["B4"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.merge_cells("B6:E6")
    sheet["B6"] = plan["workbook_subtitle"]
    sheet["B6"].font = Font(name="Aptos", size=11, color=MUTED)
    sheet["B6"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells("B8:E8")
    sheet["B8"] = "WORKBOOK REQUEST"
    sheet["B8"].font = Font(name="Aptos", size=8, bold=True, color=GOLD)
    sheet.merge_cells("B9:E11")
    sheet["B9"] = request_text
    sheet["B9"].font = Font(name="Aptos", size=10, color=INK)
    sheet["B9"].fill = PatternFill("solid", fgColor=LIGHT)
    sheet["B9"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells("B13:E13")
    sheet["B13"] = "YCHARTS REFRESH"
    sheet["B13"].font = Font(name="Aptos", size=8, bold=True, color=GOLD)
    sheet.merge_cells("B14:E16")
    sheet["B14"] = (
        "Open this workbook in desktop Excel with the YCharts Excel Add-In enabled and signed in. "
        "Use the YCharts ribbon to refresh live formulas. Historical series may populate into the "
        "reserved chart ranges after refresh."
    )
    sheet["B14"].font = Font(name="Aptos", size=10, color=INK)
    sheet["B14"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells("B18:E18")
    sheet["B18"] = "SHEETS"
    sheet["B18"].font = Font(name="Aptos", size=8, bold=True, color=GOLD)
    for row, item in enumerate(plan["sheets"], start=19):
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        sheet.cell(row, 2, f"{row - 18:02d}   {item['name']}   —   {item['title']}")
        sheet.cell(row, 2).font = Font(name="Aptos", size=10, color=NAVY)
        sheet.cell(row, 2).border = Border(bottom=Side(style="hair", color=RULE))
        sheet.row_dimensions[row].height = 22
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_area = f"A1:F{max(25, 19 + len(plan['sheets']))}"


def build_custom_workbook(plan: dict[str, object], request_text: str, output_path: Path) -> Path:
    workbook = Workbook()
    _cover(workbook, plan, request_text)
    chart_classes = {
        "area": AreaChart, "bar": BarChart, "column": BarChart,
        "doughnut": DoughnutChart, "line": LineChart, "pie": PieChart,
    }
    for item in plan["sheets"]:
        sheet = workbook.create_sheet(item["name"])
        sheet.sheet_view.showGridLines = False
        columns = item["columns"]
        _style_sheet_header(sheet, item["title"], item["subtitle"], item["source_note"], len(columns))
        for index, column in enumerate(columns, start=1):
            cell = sheet.cell(5, index, column["header"].upper())
            cell.font = Font(name="Aptos", size=8, bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="right" if column["format"] != "text" else "left", vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="medium", color=GOLD))
            sheet.column_dimensions[get_column_letter(index)].width = column["width"]
        sheet.row_dimensions[5].height = 27
        for offset, row in enumerate(item["rows"], start=6):
            for column_number, (column, cell_plan) in enumerate(zip(columns, row["cells"]), start=1):
                cell = sheet.cell(offset, column_number)
                if cell_plan["kind"] == "value":
                    value = cell_plan.get("value")
                    cell.value = "'" + value if isinstance(value, str) and value.startswith("=") else value
                    color = INK
                elif cell_plan["kind"] == "excel_formula":
                    cell.value = cell_plan["formula"]
                    color = "000000"
                else:
                    cell.value = ycharts_formula(cell_plan)
                    color = GREEN
                    cell.comment = None
                cell.font = Font(name="Aptos", size=10, color=color)
                cell.fill = PatternFill("solid", fgColor=LIGHT if offset % 2 else WHITE)
                cell.alignment = Alignment(
                    horizontal="left" if column["format"] == "text" else "right",
                    vertical="center", wrap_text=column["format"] == "text",
                )
                cell.number_format = FORMATS[column["format"]]
                cell.border = Border(bottom=Side(style="hair", color=RULE))
            sheet.row_dimensions[offset].height = 23
        last_data_row = 5 + len(item["rows"])
        sheet.auto_filter.ref = f"A5:{get_column_letter(len(columns))}{last_data_row}"
        sheet.freeze_panes = "A6"
        for column_number, column in enumerate(columns, start=1):
            if column["format"] in {"currency", "decimal", "percent", "multiple", "integer"}:
                range_string = f"{get_column_letter(column_number)}6:{get_column_letter(column_number)}{max(last_data_row, 6)}"
                sheet.conditional_formatting.add(
                    range_string,
                    CellIsRule(operator="lessThan", formula=["0"], font=Font(color=RED)),
                )
        for chart_number, chart_plan in enumerate(item["charts"]):
            chart = chart_classes[chart_plan["type"]]()
            if chart_plan["type"] == "bar":
                chart.type = "bar"
            elif chart_plan["type"] == "column":
                chart.type = "col"
            max_row = 5 + max(len(item["rows"]), chart_plan["max_rows"])
            categories = Reference(
                sheet, min_col=chart_plan["category_column"] + 1, min_row=6, max_row=max_row
            )
            series_colors = (GOLD, NAVY, "56708F", "7F8E61", "9E6B55", "6E7785")
            for series_index, series_column in enumerate(chart_plan["series_columns"]):
                data = Reference(sheet, min_col=series_column + 1, min_row=5, max_row=max_row)
                chart.add_data(data, titles_from_data=True)
                series = chart.series[-1]
                series.tx = SeriesLabel(v=str(columns[series_column]["header"]))
                if chart_plan["type"] not in {"pie", "doughnut"}:
                    series.graphicalProperties.solidFill = series_colors[series_index]
                    series.graphicalProperties.line.solidFill = series_colors[series_index]
            chart.set_categories(categories)
            chart.title = chart_plan["title"]
            chart.style = 10
            chart.height = 7.2
            chart.width = 12.4
            if len(chart_plan["series_columns"]) == 1:
                chart.legend = None
            else:
                chart.legend.position = "b"
            if hasattr(chart, "y_axis"):
                chart.y_axis.majorGridlines = None
                chart.y_axis.numFmt = FORMATS[columns[chart_plan["series_columns"][0]]["format"]]
            start_column = max(len(columns) + 2, 9)
            if len(columns) > 8:
                anchor = f"A{last_data_row + 3 + chart_number * 16}"
            else:
                anchor = f"{get_column_letter(start_column)}{5 + chart_number * 16}"
            sheet.add_chart(chart, anchor)
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.page_margins.left = .35
        sheet.page_margins.right = .35
        sheet.page_margins.top = .45
        sheet.page_margins.bottom = .45
        sheet.oddFooter.left.text = FIRM_NAME
        sheet.oddFooter.right.text = "Page &P of &N"

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    qa = OutputInspector().inspect(output_path)
    if not qa.approved:
        output_path.unlink(missing_ok=True)
        raise CustomWorkbookError("The custom workbook failed integrity checks.")
    return output_path


def build_custom_workbook_preview(plan: dict[str, object], request_text: str, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    page = landscape(letter)
    document = SimpleDocTemplate(
        str(output_path), pagesize=page, leftMargin=.45 * inch, rightMargin=.45 * inch,
        topMargin=.4 * inch, bottomMargin=.4 * inch, title=str(plan["title"]), author=FIRM_NAME,
    )
    navy = colors.HexColor("#" + NAVY)
    navy_dark = colors.HexColor("#" + NAVY_DARK)
    gold = colors.HexColor("#" + GOLD)
    muted = colors.HexColor("#" + MUTED)
    light = colors.HexColor("#" + LIGHT)
    styles = {
        "firm": ParagraphStyle("Firm", fontName="Times-Bold", fontSize=19, leading=23, textColor=colors.white),
        "title": ParagraphStyle("Title", fontName="Times-Bold", fontSize=22, leading=27, textColor=navy),
        "subtitle": ParagraphStyle("Subtitle", fontName="Helvetica", fontSize=10, leading=14, textColor=muted),
        "label": ParagraphStyle("Label", fontName="Helvetica-Bold", fontSize=7, leading=9, textColor=gold),
        "body": ParagraphStyle("Body", fontName="Helvetica", fontSize=8.5, leading=11, textColor=colors.HexColor("#" + INK)),
        "header": ParagraphStyle("Header", fontName="Helvetica-Bold", fontSize=6.6, leading=8, textColor=colors.white),
        "cell": ParagraphStyle("Cell", fontName="Helvetica", fontSize=7.2, leading=9, textColor=colors.HexColor("#" + INK)),
        "right": ParagraphStyle("Right", fontName="Helvetica", fontSize=7.2, leading=9, alignment=TA_RIGHT, textColor=colors.HexColor("#" + INK)),
    }
    story = [
        Table([[Paragraph(escape(FIRM_NAME), styles["firm"])]], colWidths=[10.1 * inch], rowHeights=[.68 * inch], style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), navy_dark), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 20),
        ])),
        Spacer(1, .28 * inch),
        Paragraph(escape(str(plan["title"])), styles["title"]),
        Paragraph(escape(str(plan["workbook_subtitle"])), styles["subtitle"]),
        Spacer(1, .22 * inch),
        Paragraph("WORKBOOK REQUEST", styles["label"]),
        Spacer(1, .07 * inch),
        Table([[Paragraph(escape(request_text), styles["body"])]], colWidths=[10.1 * inch], style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), light), ("BOX", (0, 0), (-1, -1), .5, colors.HexColor("#" + RULE)),
            ("LEFTPADDING", (0, 0), (-1, -1), 12), ("RIGHTPADDING", (0, 0), (-1, -1), 12),
            ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ])),
        Spacer(1, .22 * inch),
        Paragraph("YCHARTS REFRESH", styles["label"]),
        Paragraph(
            "Open the final workbook in desktop Excel with the YCharts add-in enabled and signed in, then refresh from the YCharts ribbon.",
            styles["body"],
        ),
    ]
    for item in plan["sheets"]:
        story.append(PageBreak())
        story.append(Paragraph(escape(str(item["title"])), styles["title"]))
        story.append(Paragraph(escape(str(item["subtitle"])), styles["subtitle"]))
        story.append(Spacer(1, .12 * inch))
        story.append(Paragraph("SOURCE  " + escape(str(item["source_note"])), styles["label"]))
        story.append(Spacer(1, .12 * inch))
        headers = [Paragraph(escape(str(column["header"]).upper()), styles["header"]) for column in item["columns"]]
        rows = [headers]
        for row in item["rows"][:18]:
            cells = []
            for column, cell in zip(item["columns"], row["cells"]):
                value = _display_value(cell)
                if len(value) > 70:
                    value = value[:67] + "..."
                cells.append(Paragraph(escape(value), styles["cell"] if column["format"] == "text" else styles["right"]))
            rows.append(cells)
        widths = [float(column["width"]) for column in item["columns"]]
        scale = (10.1 * inch) / sum(widths)
        table = Table(rows, colWidths=[width * scale for width in widths], repeatRows=1)
        commands = [
            ("BACKGROUND", (0, 0), (-1, 0), navy), ("LINEBELOW", (0, 0), (-1, 0), 1.1, gold),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5), ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]
        for row_number in range(2, len(rows), 2):
            commands.append(("BACKGROUND", (0, row_number), (-1, row_number), light))
        table.setStyle(TableStyle(commands))
        story.append(table)
        if len(item["rows"]) > 18:
            story.append(Paragraph(f"Preview shows 18 of {len(item['rows'])} rows. The final workbook contains all rows.", styles["subtitle"]))
        for chart in item["charts"]:
            story.append(Spacer(1, .09 * inch))
            story.append(Table([[
                Paragraph(
                    "<b>" + escape(str(chart["title"])) + "</b><br/>"
                    "Live chart populates in Excel after the YCharts formulas refresh.",
                    styles["body"],
                )
            ]], colWidths=[10.1 * inch], rowHeights=[.55 * inch], style=TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), light),
                ("BOX", (0, 0), (-1, -1), .5, colors.HexColor("#" + RULE)),
                ("LINEABOVE", (0, 0), (-1, 0), 1.2, gold),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
            ])))
    document.build(story)
    qa = OutputInspector().inspect(output_path)
    if not qa.approved:
        output_path.unlink(missing_ok=True)
        raise CustomWorkbookError("The custom workbook preview failed integrity checks.")
    return output_path
