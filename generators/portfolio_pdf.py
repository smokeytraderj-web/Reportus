"""Build a source-grounded portfolio PDF from a structured planning workbook."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Callable
from urllib.parse import urlsplit
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, Table, TableStyle


NAVY = colors.HexColor("#132C50")
GOLD = colors.HexColor("#E0B72E")
INK = colors.HexColor("#172434")
MUTED = colors.HexColor("#667487")
LIGHT = colors.HexColor("#F2F4F7")
RULE = colors.HexColor("#D8DEE6")
WHITE = colors.white


class PortfolioPDFError(RuntimeError):
    """The portfolio workbook could not be interpreted without guessing."""


@dataclass(frozen=True, slots=True)
class PortfolioHolding:
    ticker: str
    holding: str
    sleeve: str
    price: float
    target: float
    invest_now: float
    monthly: float
    expense_ratio: float
    valuation: str
    reason: str


@dataclass(frozen=True, slots=True)
class ResearchUpdate:
    ticker: str
    development: str
    interpretation: str
    source_ids: str


@dataclass(frozen=True, slots=True)
class SourceEntry:
    source_id: str
    item: str
    published: str
    source: str
    url: str
    support: str


@dataclass(frozen=True, slots=True)
class PortfolioWorkbookData:
    title: str
    subtitle: str
    thesis: str
    biggest_risk: str
    implementation_rule: str
    holdings: tuple[PortfolioHolding, ...]
    assumptions: tuple[tuple[str, str], ...]
    research_updates: tuple[ResearchUpdate, ...]
    sources: tuple[SourceEntry, ...]


_PORTFOLIO_ALIASES = {
    "ticker": {"ticker", "symbol"},
    "holding": {"holding", "company", "name", "security"},
    "sleeve": {"sleeve", "category", "strategy"},
    "price": {"price", "currentprice", "marketprice"},
    "target": {"target", "targetweight", "targetallocation"},
    "invest_now": {"investnow", "initialinvestment", "initialdollars"},
    "monthly": {"monthly", "monthlyinvestment", "monthlydollars"},
    "expense_ratio": {"expenseratio", "expense", "fundfee"},
    "valuation": {"valuation", "multiple"},
    "reason": {"onelinereason", "reason", "rationale", "thesis"},
}


def _normalized(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").casefold())


def _mapping(row: tuple[object, ...]) -> dict[str, int] | None:
    normalized = [_normalized(value) for value in row]
    result: dict[str, int] = {}
    for field, aliases in _PORTFOLIO_ALIASES.items():
        index = next((i for i, value in enumerate(normalized) if value in aliases), None)
        if index is None:
            return None
        result[field] = index
    return result


def _find_portfolio_sheet(workbook):
    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible":
            continue
        for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
            mapping = _mapping(row)
            if mapping is not None:
                return sheet, row_number, mapping
    raise PortfolioPDFError(
        "A portfolio table was not found. Expected headers include Ticker, Holding, Sleeve, "
        "Price, Target %, Invest now, Monthly, Expense ratio, Valuation, and One-line reason."
    )


def is_portfolio_workbook(path: Path) -> bool:
    if path.suffix.lower() not in {".xlsx", ".xlsm"} or not path.is_file():
        return False
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        try:
            _find_portfolio_sheet(workbook)
            return True
        finally:
            workbook.close()
    except Exception:
        return False


def _number(value: object, *, field: str, row: int, default: float | None = None) -> float:
    if value is None and default is not None:
        return default
    if isinstance(value, bool) or value is None:
        raise PortfolioPDFError("Portfolio row {0}: {1} must be numeric.".format(row, field))
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "")
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100
        except ValueError as exc:
            raise PortfolioPDFError("Portfolio row {0}: {1} must be numeric.".format(row, field)) from exc
    try:
        return float(text)
    except ValueError as exc:
        raise PortfolioPDFError("Portfolio row {0}: {1} must be numeric.".format(row, field)) from exc


def _next_text(sheet, label: str) -> str:
    wanted = _normalized(label)
    for row in range(1, sheet.max_row + 1):
        if _normalized(sheet.cell(row, 1).value) == wanted:
            for later in range(row + 1, min(sheet.max_row, row + 4) + 1):
                value = sheet.cell(later, 1).value
                if value is not None and str(value).strip():
                    return str(value).strip()
    return ""


def _assumptions(workbook) -> tuple[tuple[str, str], ...]:
    sheet = next((item for item in workbook.worksheets if "assumption" in item.title.casefold()), None)
    if sheet is None:
        return ()
    wanted = {
        "initialcontribution", "monthlycontribution", "basecurrency", "accountassumption",
        "benchmark", "pricetimestamp", "weightedetffee", "tax", "foreignholdings",
    }
    result: list[tuple[str, str]] = []
    for row in sheet.iter_rows(values_only=True):
        if _normalized(row[0] if row else None) in wanted and len(row) > 1 and row[1] is not None:
            result.append((str(row[0]).strip(), str(row[1]).strip()))
    return tuple(result)


def _assumption_number(items: tuple[tuple[str, str], ...], label: str, fallback: float) -> float:
    wanted = _normalized(label)
    for key, value in items:
        if _normalized(key) == wanted:
            try:
                return float(str(value).replace("$", "").replace(",", ""))
            except ValueError:
                return fallback
    return fallback


def _research_updates(workbook) -> tuple[ResearchUpdate, ...]:
    sheet = next((item for item in workbook.worksheets if "research" in item.title.casefold()), None)
    if sheet is None:
        return ()
    header = None
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [_normalized(value) for value in row]
        if "ticker" in values and "recentmaterialdevelopment" in values and "portfoliointerpretation" in values:
            header = (row_number, values)
            break
    if header is None:
        return ()
    row_number, values = header
    indexes = {name: values.index(name) for name in ("ticker", "recentmaterialdevelopment", "portfoliointerpretation")}
    source_index = values.index("sourceid") if "sourceid" in values else None
    result: list[ResearchUpdate] = []
    for row in sheet.iter_rows(min_row=row_number + 1, values_only=True):
        ticker = str(row[indexes["ticker"]] or "").strip()
        if not ticker:
            continue
        result.append(ResearchUpdate(
            ticker,
            str(row[indexes["recentmaterialdevelopment"]] or "").strip(),
            str(row[indexes["portfoliointerpretation"]] or "").strip(),
            str(row[source_index] or "").strip() if source_index is not None else "",
        ))
    return tuple(result)


def _sources(workbook) -> tuple[SourceEntry, ...]:
    sheet = next((item for item in workbook.worksheets if item.title.casefold() == "sources"), None)
    if sheet is None:
        return ()
    header = None
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [_normalized(value) for value in row]
        if all(value in values for value in ("id", "item", "source", "url", "whatitsupports")):
            header = (row_number, values)
            break
    if header is None:
        return ()
    row_number, values = header
    def index(name: str) -> int | None:
        return values.index(name) if name in values else None
    indexes = {name: index(name) for name in ("id", "item", "asofpublished", "source", "url", "whatitsupports")}
    result: list[SourceEntry] = []
    for row in sheet.iter_rows(min_row=row_number + 1, values_only=True):
        source_id = str(row[indexes["id"]] or "").strip() if indexes["id"] is not None else ""
        if not source_id:
            continue
        def text_at(name: str) -> str:
            item_index = indexes[name]
            return str(row[item_index] or "").strip() if item_index is not None else ""
        result.append(SourceEntry(
            source_id, text_at("item"), text_at("asofpublished"), text_at("source"),
            text_at("url"), text_at("whatitsupports"),
        ))
    return tuple(result)


def load_portfolio_workbook(path: Path) -> PortfolioWorkbookData:
    workbook = load_workbook(path, read_only=False, data_only=True, keep_links=False)
    try:
        sheet, header_row, mapping = _find_portfolio_sheet(workbook)
        assumptions = _assumptions(workbook)
        initial = _assumption_number(assumptions, "Initial contribution", 100.0)
        monthly_contribution = _assumption_number(assumptions, "Monthly contribution", 100.0)
        holdings: list[PortfolioHolding] = []
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
        ):
            def at(field: str):
                index = mapping[field]
                return row[index] if index < len(row) else None
            ticker = str(at("ticker") or "").strip()
            holding = str(at("holding") or "").strip()
            if _normalized(ticker) == "total":
                break
            if not ticker and not holding:
                continue
            if not ticker or not holding:
                raise PortfolioPDFError(
                    "Portfolio row {0}: ticker and holding are both required.".format(row_number)
                )
            target = _number(at("target"), field="target percentage", row=row_number)
            holdings.append(PortfolioHolding(
                ticker=ticker,
                holding=holding,
                sleeve=str(at("sleeve") or "").strip(),
                price=_number(at("price"), field="price", row=row_number),
                target=target,
                invest_now=_number(at("invest_now"), field="invest now", row=row_number, default=target * initial),
                monthly=_number(at("monthly"), field="monthly contribution", row=row_number, default=target * monthly_contribution),
                expense_ratio=_number(at("expense_ratio"), field="expense ratio", row=row_number, default=0.0),
                valuation=str(at("valuation") or "").strip(),
                reason=str(at("reason") or "").strip(),
            ))
        if not holdings:
            raise PortfolioPDFError("No portfolio holdings were found below the detected headers.")
        total_target = sum(item.target for item in holdings)
        if abs(total_target - 1.0) > 0.005:
            raise PortfolioPDFError(
                "Portfolio target weights total {0:.2%}; they must total 100%.".format(total_target)
            )
        title = str(sheet["A1"].value or "Portfolio Review").strip()
        subtitle = str(sheet["A2"].value or "").strip()
        return PortfolioWorkbookData(
            title=title,
            subtitle=subtitle,
            thesis=str(sheet["A5"].value or "").strip(),
            biggest_risk=_next_text(sheet, "Biggest risk"),
            implementation_rule=next(
                (str(sheet.cell(row, 1).value).strip() for row in range(1, sheet.max_row + 1)
                 if str(sheet.cell(row, 1).value or "").strip().casefold().startswith("implementation rule")),
                "",
            ),
            holdings=tuple(holdings),
            assumptions=assumptions,
            research_updates=_research_updates(workbook),
            sources=_sources(workbook),
        )
    finally:
        workbook.close()


def _page_size(reference_pdf: Path) -> tuple[float, float]:
    reader = PdfReader(reference_pdf)
    if not reader.pages:
        raise PortfolioPDFError("The reference PDF has no pages.")
    page = reader.pages[0]
    width, height = float(page.mediabox.width), float(page.mediabox.height)
    rotation = int(page.get("/Rotate", 0) or 0) % 360
    if rotation in (90, 270):
        width, height = height, width
    if width <= 0 or height <= 0:
        raise PortfolioPDFError("The reference PDF page size is invalid.")
    return width, height


def _wrapped_lines(text: str, font: str, size: float, width: float) -> list[str]:
    words = str(text or "").split()
    lines: list[str] = []
    current = ""
    for word in words:
        candidate = word if not current else current + " " + word
        if stringWidth(candidate, font, size) <= width or not current:
            current = candidate
        else:
            lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def _draw_wrapped(c, text: str, x: float, y: float, width: float, *, font="Helvetica", size=10,
                  leading=14, color=INK, max_lines: int | None = None) -> float:
    lines = _wrapped_lines(text, font, size, width)
    if max_lines is not None:
        lines = lines[:max_lines]
    c.setFillColor(color)
    c.setFont(font, size)
    for line in lines:
        c.drawString(x, y, line)
        y -= leading
    return y


def _paragraph(text: object, style: ParagraphStyle) -> Paragraph:
    return Paragraph(escape(str(text or "")), style)


def _money(value: float) -> str:
    return "${0:,.2f}".format(value)


def _percent(value: float, decimals: int = 1) -> str:
    return ("{0:." + str(decimals) + "%}").format(value)


class _PortfolioRenderer:
    def __init__(self, output_path: Path, page_size: tuple[float, float], source_label: str):
        self.output_path = output_path
        self.width, self.height = page_size
        self.margin = max(34.0, self.width * 0.055)
        self.content_width = self.width - 2 * self.margin
        self.c = canvas.Canvas(str(output_path), pagesize=page_size, pageCompression=1)
        self.page_number = 0
        self.source_label = source_label
        self.body = ParagraphStyle("body", fontName="Helvetica", fontSize=8, leading=10, textColor=INK)
        self.small = ParagraphStyle("small", fontName="Helvetica", fontSize=6.5, leading=8, textColor=MUTED)
        self.header = ParagraphStyle("header", fontName="Helvetica-Bold", fontSize=6.5, leading=8, textColor=WHITE)

    def _new_page(self, title: str, eyebrow: str = "PORTFOLIO REPORT") -> float:
        if self.page_number:
            self.c.showPage()
        self.page_number += 1
        self.c.setFillColor(WHITE)
        self.c.rect(0, 0, self.width, self.height, fill=1, stroke=0)
        self.c.setFillColor(NAVY)
        self.c.rect(0, self.height - 68, self.width, 68, fill=1, stroke=0)
        self.c.setFillColor(GOLD)
        self.c.setFont("Helvetica-Bold", 7)
        self.c.drawString(self.margin, self.height - 22, eyebrow.upper())
        self.c.setFillColor(WHITE)
        self.c.setFont("Helvetica-Bold", 20)
        self.c.drawString(self.margin, self.height - 50, title)
        self._footer()
        return self.height - 92

    def _footer(self) -> None:
        self.c.setStrokeColor(RULE)
        self.c.line(self.margin, 24, self.width - self.margin, 24)
        self.c.setFillColor(MUTED)
        self.c.setFont("Helvetica", 6.5)
        self.c.drawString(self.margin, 12, "Source: " + self.source_label)
        self.c.drawRightString(self.width - self.margin, 12, "Page {0}".format(self.page_number))

    def cover(self, data: PortfolioWorkbookData, client_name: str, report_title: str,
              period_label: str) -> None:
        self.page_number = 1
        self.c.setFillColor(NAVY)
        self.c.rect(0, 0, self.width, self.height, fill=1, stroke=0)
        self.c.setFillColor(GOLD)
        self.c.setFont("Helvetica-Bold", 8)
        self.c.drawString(self.margin, self.height - 70, "PORTFOLIO CONSULTING  ·  LONG-TERM ALLOCATION")
        y = self.height - 120
        y = _draw_wrapped(
            self.c, report_title or data.title, self.margin, y, self.content_width * 0.85,
            font="Helvetica-Bold", size=30, leading=34, color=WHITE, max_lines=3,
        )
        self.c.setFillColor(GOLD)
        self.c.rect(self.margin, y - 8, 150, 3, fill=1, stroke=0)
        y -= 45
        y = _draw_wrapped(
            self.c, data.title, self.margin, y, self.content_width * 0.82,
            size=11, leading=15, color=colors.HexColor("#D4DCE8"), max_lines=2,
        )
        meta_y = 80
        for x, label, value in (
            (self.margin, "PREPARED FOR", client_name or "Client review"),
            (self.margin + self.content_width * 0.38, "REPORT PERIOD", period_label or "Current review"),
            (self.margin + self.content_width * 0.68, "HOLDINGS", str(len(data.holdings))),
        ):
            self.c.setFillColor(colors.HexColor("#91A2BA"))
            self.c.setFont("Helvetica-Bold", 6.5)
            self.c.drawString(x, meta_y + 22, label)
            self.c.setFillColor(WHITE)
            self.c.setFont("Helvetica-Bold", 10)
            self.c.drawString(x, meta_y, value[:48])

    def overview(self, data: PortfolioWorkbookData) -> None:
        y = self._new_page("Portfolio at a glance")
        initial = sum(item.invest_now for item in data.holdings)
        monthly = sum(item.monthly for item in data.holdings)
        weighted_fee = sum(item.target * item.expense_ratio for item in data.holdings)
        metrics = (
            ("INITIAL INVESTMENT", _money(initial)),
            ("MONTHLY CONTRIBUTION", _money(monthly)),
            ("TARGET WEIGHT", _percent(sum(item.target for item in data.holdings), 0)),
            ("WEIGHTED ETF FEE", _percent(weighted_fee, 2)),
        )
        gap = 10
        box_width = (self.content_width - gap * 3) / 4
        for index, (label, value) in enumerate(metrics):
            x = self.margin + index * (box_width + gap)
            self.c.setFillColor(LIGHT)
            self.c.roundRect(x, y - 70, box_width, 58, 5, fill=1, stroke=0)
            self.c.setFillColor(MUTED)
            self.c.setFont("Helvetica-Bold", 6.5)
            self.c.drawString(x + 10, y - 31, label)
            self.c.setFillColor(NAVY)
            self.c.setFont("Helvetica-Bold", 16)
            self.c.drawString(x + 10, y - 55, value)
        y -= 100
        self.c.setFillColor(NAVY)
        self.c.setFont("Helvetica-Bold", 10)
        self.c.drawString(self.margin, y, "INVESTMENT THESIS")
        _draw_wrapped(
            self.c, data.thesis or "No thesis was supplied.", self.margin, y - 24,
            self.content_width, size=9, leading=14, color=INK, max_lines=9,
        )

    def allocation(self, data: PortfolioWorkbookData) -> None:
        chunks = [data.holdings[index:index + 12] for index in range(0, len(data.holdings), 12)]
        for page_index, holdings in enumerate(chunks, start=1):
            title = "Target allocation" if page_index == 1 else "Target allocation (continued)"
            y = self._new_page(title)
            headers = ("Ticker", "Holding", "Sleeve", "Price", "Target", "Now", "Monthly", "Valuation / rationale")
            rows = [[_paragraph(value, self.header) for value in headers]]
            for item in holdings:
                detail = "<b>{0}</b><br/>{1}".format(escape(item.valuation), escape(item.reason))
                rows.append([
                    _paragraph(item.ticker, self.body), _paragraph(item.holding, self.body),
                    _paragraph(item.sleeve, self.small), _paragraph(_money(item.price), self.body),
                    _paragraph(_percent(item.target, 0), self.body), _paragraph(_money(item.invest_now), self.body),
                    _paragraph(_money(item.monthly), self.body), Paragraph(detail, self.small),
                ])
            fractions = (0.065, 0.15, 0.11, 0.075, 0.065, 0.075, 0.075, 0.385)
            table = Table(rows, colWidths=[self.content_width * value for value in fractions], repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), NAVY),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT]),
                ("LINEBELOW", (0, 1), (-1, -1), 0.25, RULE),
            ]))
            _, table_height = table.wrap(self.content_width, y - 38)
            table.drawOn(self.c, self.margin, y - table_height)

    def risks(self, data: PortfolioWorkbookData) -> None:
        y = self._new_page("Risk, assumptions and implementation")
        for title, body in (
            ("BIGGEST RISK", data.biggest_risk),
            ("IMPLEMENTATION RULE", data.implementation_rule),
        ):
            self.c.setFillColor(NAVY)
            self.c.setFont("Helvetica-Bold", 9)
            self.c.drawString(self.margin, y, title)
            y = _draw_wrapped(
                self.c, body or "Not supplied.", self.margin, y - 18, self.content_width,
                size=8, leading=11, color=INK, max_lines=7,
            ) - 20
        self.c.setFillColor(NAVY)
        self.c.setFont("Helvetica-Bold", 9)
        self.c.drawString(self.margin, y, "KEY ASSUMPTIONS")
        y -= 16
        for label, value in data.assumptions[:9]:
            self.c.setFillColor(MUTED)
            self.c.setFont("Helvetica-Bold", 7)
            self.c.drawString(self.margin, y, label.upper())
            y = _draw_wrapped(
                self.c, value, self.margin + self.content_width * 0.24, y,
                self.content_width * 0.74, size=7.5, leading=10, color=INK, max_lines=2,
            ) - 5

    def research(self, data: PortfolioWorkbookData) -> None:
        if not data.research_updates:
            return
        chunks = [data.research_updates[index:index + 8] for index in range(0, len(data.research_updates), 8)]
        for page_index, updates in enumerate(chunks, start=1):
            title = "Recent research checks" if page_index == 1 else "Recent research checks (continued)"
            y = self._new_page(title, "SOURCE-GROUNDED REVIEW")
            rows = [[_paragraph(value, self.header) for value in ("Ticker", "Recent development", "Portfolio interpretation")]]
            for item in updates:
                interpretation = item.interpretation
                if item.source_ids:
                    interpretation += "  [" + item.source_ids + "]"
                rows.append([
                    _paragraph(item.ticker, self.body), _paragraph(item.development, self.small),
                    _paragraph(interpretation, self.small),
                ])
            table = Table(rows, colWidths=[self.content_width * 0.08, self.content_width * 0.46, self.content_width * 0.46])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), NAVY),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT]),
                ("LINEBELOW", (0, 1), (-1, -1), 0.25, RULE),
            ]))
            _, table_height = table.wrap(self.content_width, y - 35)
            table.drawOn(self.c, self.margin, y - table_height)

    def sources(self, data: PortfolioWorkbookData) -> None:
        if not data.sources:
            return
        chunks = [data.sources[index:index + 10] for index in range(0, len(data.sources), 10)]
        for page_index, sources in enumerate(chunks, start=1):
            title = "Sources and data audit" if page_index == 1 else "Sources and data audit (continued)"
            y = self._new_page(title, "AUDIT TRAIL")
            rows = [[_paragraph(value, self.header) for value in ("ID", "Item", "Published", "Source", "What it supports")]]
            for item in sources:
                domain = urlsplit(item.url).netloc.removeprefix("www.") if item.url else ""
                source = item.source + (" · " + domain if domain else "")
                rows.append([
                    _paragraph(item.source_id, self.body), _paragraph(item.item, self.small),
                    _paragraph(item.published, self.small), _paragraph(source, self.small),
                    _paragraph(item.support, self.small),
                ])
            fractions = (0.06, 0.19, 0.12, 0.22, 0.41)
            table = Table(rows, colWidths=[self.content_width * value for value in fractions])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), NAVY),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT]),
                ("LINEBELOW", (0, 1), (-1, -1), 0.25, RULE),
            ]))
            _, table_height = table.wrap(self.content_width, y - 35)
            table.drawOn(self.c, self.margin, y - table_height)

    def save(self) -> Path:
        self.c.save()
        reader = PdfReader(self.output_path)
        if not reader.pages:
            raise PortfolioPDFError("The generated portfolio PDF contains no pages.")
        return self.output_path


def build_portfolio_workbook_pdf(
    workbook_path: Path,
    reference_pdf: Path,
    output_path: Path,
    *,
    client_name: str,
    report_title: str,
    period_label: str,
    source_label: str,
    converter: Callable[[Path, Path], Path] | None = None,
) -> Path:
    """Build a portfolio report through the official branded client-deck system."""

    data = load_portfolio_workbook(workbook_path)
    _page_size(reference_pdf)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    intermediate = output_path.with_suffix(".pptx")
    from generators.portfolio_client_deck import build_portfolio_client_deck
    from services.conversion import convert_pptx_to_pdf

    build_portfolio_client_deck(
        data,
        intermediate,
        client_name=client_name,
        report_title=report_title,
        period_label=period_label,
        source_label=source_label,
    )
    (converter or convert_pptx_to_pdf)(intermediate, output_path)
    reader = PdfReader(output_path)
    if not reader.pages:
        raise PortfolioPDFError("The generated portfolio PDF contains no pages.")
    return output_path
