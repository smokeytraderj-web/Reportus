"""Generate a branded stock-review PDF from a validated Excel workbook.

The layout comes from the installed ``template-pdf-report`` skill.  This module
owns Reportus-specific safety rules: no inferred investment calls, no hardcoded
client paths, escaped workbook text, portable browser discovery, and verified
temporary rendering before the final file is copied into place.
"""

from __future__ import annotations

import datetime as dt
import html
import importlib.util
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from config.settings import PROJECT_ROOT
from quality.output_qa import OutputInspector


FIRM_NAME = "Gottfried & Somberg Wealth Management"
RATING_WORDS = frozenset(
    {"buy", "sell", "hold", "pass", "add", "trim", "build", "strong",
     "reiterate", "overweight", "underweight", "neutral"}
)
INTERNAL_SOURCE_NAMES = frozenset({"dave", "matt", "nick"})


class ExcelToPDFError(RuntimeError):
    """A safe, user-facing Excel-to-PDF build failure."""


@dataclass(frozen=True, slots=True)
class ReviewDefinition:
    sheet_name: str
    theme: str


@dataclass(frozen=True, slots=True)
class StockReviewConfig:
    client_name: str
    report_title: str
    period_label: str
    reviews: tuple[ReviewDefinition, ...]
    source_label: str = "Uploaded workbook"
    firm_name: str = FIRM_NAME
    rows_per_page: int = 13


@dataclass(frozen=True, slots=True)
class StockRow:
    company: str
    ticker: str
    rating: str
    price: float
    total_return: float
    versus_benchmark: float
    notes: str


@dataclass(frozen=True, slots=True)
class StockSection:
    name: str
    rows: tuple[StockRow, ...]


@dataclass(frozen=True, slots=True)
class StockReview:
    number: int
    theme: str
    recommendation_date: dt.date
    as_of_date: dt.date
    benchmark_return: float
    sections: tuple[StockSection, ...]


def _load_skill_css() -> str:
    reference = (
        PROJECT_ROOT / "skills" / "template-pdf-report" / "references" /
        "build_stock_picks.py"
    )
    spec = importlib.util.spec_from_file_location("reportus_stock_pdf_reference", reference)
    if spec is None or spec.loader is None:
        raise ExcelToPDFError("The Excel-to-PDF style resource is unavailable.")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return str(module.CSS)


def _date(value: object, *, field: str, sheet: str) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    raise ExcelToPDFError(f"{sheet}: {field} must be a valid Excel date.")


def _number(value: object, *, field: str, sheet: str, row: int | None = None) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        where = f" row {row}" if row is not None else ""
        raise ExcelToPDFError(f"{sheet}{where}: {field} must be numeric.")
    return float(value)


def _split_rating(note: str) -> tuple[str | None, str]:
    """Extract an explicit leading rating without changing the source meaning."""

    note = note.split(">>", 1)[0].strip()
    index = note.find(" - ")
    if index == -1 or index > 40:
        return None, note
    lead = re.sub(r"\s*\([^)]*\)", "", note[:index]).strip()
    tokens = tuple(token for token in re.split(r"[/\s]+", lead.lower()) if token)
    if tokens and all(token in RATING_WORDS for token in tokens):
        return lead, note[index + 3:].strip()
    return None, note


def load_stock_reviews(workbook_path: Path, config: StockReviewConfig) -> tuple[StockReview, ...]:
    """Read the skill's workbook contract and reject missing or ambiguous values."""

    if not workbook_path.is_file():
        raise ExcelToPDFError("The workbook is missing.")
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ExcelToPDFError("This report requires an .xlsx or .xlsm workbook.")
    if not config.client_name.strip():
        raise ExcelToPDFError("Client name is required.")
    if not config.reviews:
        raise ExcelToPDFError("At least one review sheet must be configured.")
    if not 5 <= config.rows_per_page <= 20:
        raise ExcelToPDFError("Rows per page must be between 5 and 20.")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True, keep_links=False)
    try:
        missing = [item.sheet_name for item in config.reviews if item.sheet_name not in workbook.sheetnames]
        if missing:
            raise ExcelToPDFError("Missing required worksheet(s): " + ", ".join(missing))

        reviews: list[StockReview] = []
        for number, definition in enumerate(config.reviews, start=1):
            sheet = workbook[definition.sheet_name]
            as_of: object = None
            recommendation_date: object = None
            benchmark: object = None
            sections: list[tuple[str, list[StockRow]]] = []

            for row_number, raw in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = list(raw) + [None] * max(0, 10 - len(raw))
                label_value = values[1]
                if label_value is None:
                    continue
                label = str(label_value).strip()
                if label.startswith("As-Of Date"):
                    as_of = values[2]
                    continue
                if label.startswith("Default Rec. Date"):
                    recommendation_date = values[2]
                    continue
                if label.startswith("S&P 500 Total Return (benchmark)"):
                    benchmark = values[5]
                    continue
                if label == "Company Name" or label.endswith("(live YCharts)"):
                    continue

                ticker_value, marker = values[2], values[3]
                if ticker_value is None and marker is None:
                    sections.append((label, []))
                    continue
                if not sections:
                    raise ExcelToPDFError(
                        f"{definition.sheet_name} row {row_number}: add a section heading before data rows."
                    )
                if ticker_value is None:
                    raise ExcelToPDFError(f"{definition.sheet_name} row {row_number}: ticker is required.")

                raw_note = str(values[7] or "").strip()
                rating, notes = _split_rating(raw_note)
                if rating is None:
                    raise ExcelToPDFError(
                        f"{definition.sheet_name} row {row_number}: an explicit rating is required for "
                        f"{str(ticker_value).strip()}. Add it to the note as 'Buy - ...', 'Hold - ...', or similar."
                    )
                source_firm = str(values[9] or "").strip()
                if source_firm and source_firm.casefold() not in INTERNAL_SOURCE_NAMES:
                    notes = f"{notes.rstrip()} (Source: {source_firm})"

                sections[-1][1].append(
                    StockRow(
                        company=label,
                        ticker=str(ticker_value).strip(),
                        rating=rating,
                        price=_number(values[8], field="price at recommendation", sheet=definition.sheet_name, row=row_number),
                        total_return=_number(values[5], field="total return", sheet=definition.sheet_name, row=row_number),
                        versus_benchmark=_number(values[6], field="return versus benchmark", sheet=definition.sheet_name, row=row_number),
                        notes=notes,
                    )
                )

            if as_of is None or recommendation_date is None or benchmark is None:
                raise ExcelToPDFError(
                    f"{definition.sheet_name}: As-Of Date, Default Rec. Date, and benchmark are required."
                )
            nonempty = tuple(StockSection(name, tuple(rows)) for name, rows in sections if rows)
            if not nonempty:
                raise ExcelToPDFError(f"{definition.sheet_name}: no report rows were found.")
            reviews.append(
                StockReview(
                    number=number,
                    theme=definition.theme,
                    recommendation_date=_date(recommendation_date, field="Default Rec. Date", sheet=definition.sheet_name),
                    as_of_date=_date(as_of, field="As-Of Date", sheet=definition.sheet_name),
                    benchmark_return=_number(benchmark, field="benchmark", sheet=definition.sheet_name),
                    sections=nonempty,
                )
            )
        return tuple(reviews)
    finally:
        workbook.close()


def _fmt_date(value: dt.date) -> str:
    return f"{value.strftime('%B')} {value.day}, {value.year}"


def _pct(value: float) -> str:
    return f"{'+' if value >= 0 else ''}{value * 100:.2f}%"


def _money(value: float) -> str:
    return f"${value:,.2f}"


def _tone(value: float) -> str:
    return "pos" if value >= 0 else "neg"


def _paginate(review: StockReview, rows_per_page: int) -> tuple[tuple[StockSection, ...], ...]:
    pages: list[list[StockSection]] = []
    current: list[StockSection] = []
    used = 0
    for section in review.sections:
        remaining = list(section.rows)
        continued = False
        while remaining:
            capacity = rows_per_page - used
            if capacity <= 0:
                pages.append(current)
                current, used = [], 0
                capacity = rows_per_page
            chunk, remaining = remaining[:capacity], remaining[capacity:]
            name = section.name + (" (CONT'D)" if continued else "")
            current.append(StockSection(name, tuple(chunk)))
            used += len(chunk)
            continued = True
            if remaining:
                pages.append(current)
                current, used = [], 0
    if current:
        pages.append(current)
    return tuple(tuple(page) for page in pages)


def _footer(config: StockReviewConfig, page: int, total: int) -> str:
    return (
        '<div class="foot"><span class="firm">{firm}</span>'
        '<span>Source: {source}</span><span>Page {page} of {total}</span></div>'
    ).format(
        firm=html.escape(config.firm_name), source=html.escape(config.source_label),
        page=page, total=total,
    )


def build_stock_review_html(reviews: tuple[StockReview, ...], config: StockReviewConfig) -> str:
    """Build fixed-page HTML after all page ranges have been computed."""

    if not reviews:
        raise ExcelToPDFError("No review data was supplied.")
    paged = tuple((review, _paginate(review, config.rows_per_page)) for review in reviews)
    total_pages = 2 + sum(1 + len(pages) for _, pages in paged)
    page_ranges: dict[int, tuple[int, int]] = {}
    cursor = 2
    for review, pages in paged:
        first = cursor + 1
        cursor += 1 + len(pages)
        page_ranges[review.number] = (first, cursor)

    latest_as_of = max(review.as_of_date for review in reviews)
    parts = [
        '<div class="page dark">'
        '<div class="eyebrow cover-eyebrow">PRIVATE CLIENT REVIEW</div>'
        f'<h1 class="cover-h1">{html.escape(config.client_name)}</h1>'
        f'<div class="cover-sub">{html.escape(config.report_title)} &nbsp;·&nbsp; {html.escape(config.period_label)}</div>'
        '<div class="cover-rule"></div>'
        '<div class="cover-desc">A sourced record of investment recommendations, theses, and performance versus the S&amp;P 500 Total Return Index.</div>'
        f'<div class="cover-foot">Prepared for Client Review &nbsp;|&nbsp; Data as of {_fmt_date(latest_as_of)} &nbsp;|&nbsp; Source: {html.escape(config.source_label)}</div>'
        '</div>'
    ]

    toc_rows = []
    for review in reviews:
        first, last = page_ranges[review.number]
        span = f"Page {first}" if first == last else f"Pages {first}&ndash;{last}"
        toc_rows.append(
            f'<tr><td class="appt">Portfolio Review {review.number}</td>'
            f'<td class="theme">{html.escape(review.theme)}</td>'
            f'<td class="pages r">{span}</td></tr>'
        )
    parts.append(
        '<div class="page"><div class="band-flat"><h1>Contents</h1></div>'
        '<table class="toc"><colgroup><col class="t1"><col class="t2"><col class="t4"></colgroup>'
        '<thead><tr><th>Portfolio Review</th><th>Key Themes</th><th class="r">Pages</th></tr></thead>'
        f'<tbody>{"".join(toc_rows)}</tbody></table></div>'
    )

    page_number = 2
    rendered_rows = 0
    for review, pages in paged:
        page_number += 1
        section_items = "".join(
            f'<li>&mdash;&nbsp; {html.escape(section.name)}</li>' for section in review.sections
        )
        parts.append(
            '<div class="page dark">'
            f'<h1 class="div-h1">Portfolio Review {review.number}</h1>'
            f'<div class="div-sub">Recommendation date {_fmt_date(review.recommendation_date)} &nbsp;·&nbsp; measured through {_fmt_date(review.as_of_date)}</div>'
            '<div class="spbox"><div class="lbl">S&amp;P 500 TOTAL RETURN</div>'
            f'<div class="val {_tone(review.benchmark_return)}">{_pct(review.benchmark_return)}</div></div>'
            '<div class="eyebrow div-listhead">IN THIS SECTION</div>'
            f'<ul class="div-list">{section_items}</ul>'
            f'<div class="div-foot"><span>{html.escape(config.firm_name)}</span><span>Page {page_number} of {total_pages}</span></div>'
            '</div>'
        )
        for part_number, sections in enumerate(pages, start=1):
            page_number += 1
            table_rows: list[str] = []
            for section in sections:
                table_rows.append(f'<tr class="sect"><td colspan="6">{html.escape(section.name)}</td></tr>')
                for row in section.rows:
                    rendered_rows += 1
                    table_rows.append(
                        '<tr>'
                        f'<td class="tick">{html.escape(row.ticker)}<div class="co">{html.escape(row.company)}</div></td>'
                        f'<td class="rate"><span class="pill">{html.escape(row.rating)}</span></td>'
                        f'<td class="num">{_money(row.price)}</td>'
                        f'<td class="num b {_tone(row.total_return)}">{_pct(row.total_return)}</td>'
                        f'<td class="num b {_tone(row.versus_benchmark)}">{_pct(row.versus_benchmark)}</td>'
                        f'<td class="note">{html.escape(row.notes)}</td>'
                        '</tr>'
                    )
            parts.append(
                '<div class="page"><div class="band">'
                f'<h1>Portfolio Review {review.number} ({part_number} of {len(pages)})</h1>'
                '<div class="band-meta">'
                f'<div>Total return from <b>{_fmt_date(review.recommendation_date)}</b> through <b>{_fmt_date(review.as_of_date)}</b></div>'
                f'<div>S&amp;P 500 Total Return <b>{_pct(review.benchmark_return)}</b></div></div></div>'
                '<div class="tbl-wrap"><table class="tbl">'
                '<colgroup><col class="c1"><col class="c2"><col class="c3"><col class="c4"><col class="c5"><col class="c6"></colgroup>'
                '<thead><tr><th>Name</th><th>Rating</th><th class="r">Price at Rec.</th><th class="r">Return</th><th class="r">vs. S&amp;P</th><th>Notes</th></tr></thead>'
                f'<tbody>{"".join(table_rows)}</tbody></table></div>'
                f'{_footer(config, page_number, total_pages)}</div>'
            )

    source_rows = sum(len(section.rows) for review in reviews for section in review.sections)
    if rendered_rows != source_rows:
        raise ExcelToPDFError("Internal row-count check failed; the report was not rendered.")
    css = _load_skill_css()
    title = html.escape(f"{config.client_name} - {config.report_title}")
    return f'<!DOCTYPE html><html><head><meta charset="utf-8"><title>{title}</title><style>{css}</style></head><body>{"".join(parts)}</body></html>'


def find_chromium(explicit: Path | None = None) -> Path:
    """Locate Chrome/Chromium on Windows, macOS, or Linux."""

    candidates: Iterable[Path]
    if explicit is not None:
        candidates = (explicit,)
    else:
        names = ("google-chrome", "google-chrome-stable", "chromium", "chromium-browser", "chrome")
        discovered = [Path(found) for name in names if (found := shutil.which(name))]
        candidates = (
            *discovered,
            Path("C:/Program Files/Google/Chrome/Application/chrome.exe"),
            Path("C:/Program Files (x86)/Google/Chrome/Application/chrome.exe"),
            Path.home() / "AppData/Local/Google/Chrome/Application/chrome.exe",
            Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
            Path("/Applications/Chromium.app/Contents/MacOS/Chromium"),
        )
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    raise ExcelToPDFError("Google Chrome or Chromium is required to build this PDF.")


def render_html_to_pdf(document: str, destination: Path, *, browser_path: Path | None = None,
                       keep_html: Path | None = None) -> Path:
    """Render through an isolated temp file and accept only a verified PDF."""

    browser = find_chromium(browser_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    if keep_html is not None:
        keep_html.parent.mkdir(parents=True, exist_ok=True)
        keep_html.write_text(document, encoding="utf-8")
    with tempfile.TemporaryDirectory(prefix="reportus-pdf-") as temporary:
        temp_dir = Path(temporary)
        html_path = temp_dir / "report.html"
        pdf_path = temp_dir / "report.pdf"
        html_path.write_text(document, encoding="utf-8")
        completed = subprocess.run(
            [
                str(browser), "--headless", "--disable-gpu", "--no-pdf-header-footer",
                "--no-margins", f"--print-to-pdf={pdf_path}", html_path.resolve().as_uri(),
            ],
            capture_output=True,
            text=True,
            timeout=180,
            check=False,
        )
        if completed.returncode != 0 or not pdf_path.is_file() or pdf_path.stat().st_size < 100:
            raise ExcelToPDFError("Chrome did not produce a usable PDF.")
        qa = OutputInspector().inspect(pdf_path)
        if not qa.approved:
            raise ExcelToPDFError("The generated PDF failed final integrity checks.")
        try:
            shutil.copy2(pdf_path, destination)
        except OSError as exc:
            raise ExcelToPDFError(
                "The destination PDF is open or unavailable. Close it and try again."
            ) from exc
    return destination


def build_stock_review_pdf(workbook_path: Path, output_path: Path, config: StockReviewConfig,
                           *, browser_path: Path | None = None, keep_html: Path | None = None) -> Path:
    """Build a stock-review PDF from the workbook in one validated call."""

    reviews = load_stock_reviews(workbook_path, config)
    document = build_stock_review_html(reviews, config)
    return render_html_to_pdf(
        document, output_path, browser_path=browser_path, keep_html=keep_html
    )
