"""Build the current clean/minimal GSWM holdings workbook and PDF snapshot."""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from quality.output_qa import OutputInspector


NAVY = "172D4F"
GOLD = "B49A58"
INK = "1B2532"
GRAY = "6E7785"
LIGHT = "F5F6F8"
WHITE = "FFFFFF"
FIRM_NAME = "Gottfried & Somberg Wealth Management"


class WorkbookBuildError(RuntimeError):
    pass


def _pdf_color(value: str):
    return colors.HexColor(f"#{value}")


@dataclass(frozen=True, slots=True)
class Holding:
    company: str
    symbol: str
    quantity: float
    price: float
    value: float
    percent_assets: float


@dataclass(frozen=True, slots=True)
class HoldingsWorkbookConfig:
    report_label: str
    source_label: str = "Uploaded source data"
    firm_label: str = FIRM_NAME
    sheet_name: str = "Holdings Summary"


_ALIASES = {
    "company": {"description", "company", "companyname", "holding", "securitydescription", "name"},
    "symbol": {"symbol", "ticker", "tickersymbol"},
    "quantity": {"quantity", "qty", "shares", "units"},
    "price": {"price", "currentprice", "marketprice"},
    "value": {"value", "marketvalue", "currentvalue"},
    "percent_assets": {"ofassets", "percentofassets", "pctofassets", "weight", "portfolioweight"},
}


def _normalized(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").casefold())


def _columns(values: tuple[object, ...] | list[object]) -> dict[str, int] | None:
    normalized = [_normalized(value) for value in values]
    result: dict[str, int] = {}
    for field, aliases in _ALIASES.items():
        index = next((i for i, value in enumerate(normalized) if value in aliases), None)
        if index is None:
            return None
        result[field] = index
    return result


def _number(value: object, *, field: str, row: int) -> float:
    if isinstance(value, bool) or value is None:
        raise WorkbookBuildError(f"Row {row}: {field} is required and must be numeric.")
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "")
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    try:
        parsed = float(text)
    except ValueError as exc:
        raise WorkbookBuildError(f"Row {row}: {field} must be numeric.") from exc
    return -parsed if negative else parsed


def _percent(value: object, *, row: int) -> float:
    if isinstance(value, str) and value.strip().endswith("%"):
        return _number(value.strip()[:-1], field="% of assets", row=row) / 100
    parsed = _number(value, field="% of assets", row=row)
    if abs(parsed) > 1:
        raise WorkbookBuildError(
            f"Row {row}: % of assets is ambiguous. Use an Excel percentage or include the % sign."
        )
    return parsed


def _holding(values: list[object], mapping: dict[str, int], row_number: int) -> Holding | None:
    def at(field: str) -> object:
        index = mapping[field]
        return values[index] if index < len(values) else None

    company = str(at("company") or "").strip()
    symbol = str(at("symbol") or "").strip()
    if not company and not symbol:
        return None
    if not symbol and company.casefold().startswith("total"):
        return None
    if not company or not symbol:
        raise WorkbookBuildError(f"Row {row_number}: company and symbol are both required.")
    return Holding(
        company=company,
        symbol=symbol,
        quantity=_number(at("quantity"), field="quantity", row=row_number),
        price=_number(at("price"), field="price", row=row_number),
        value=_number(at("value"), field="value", row=row_number),
        percent_assets=_percent(at("percent_assets"), row=row_number),
    )


def _from_rows(rows: list[list[object]]) -> tuple[Holding, ...]:
    header_index = None
    mapping = None
    for index, row in enumerate(rows[:50]):
        candidate = _columns(row)
        if candidate is not None:
            header_index, mapping = index, candidate
            break
    if header_index is None or mapping is None:
        raise WorkbookBuildError(
            "Required columns were not found: Description, Symbol, Quantity, Price, Value, and % of Assets."
        )
    holdings = tuple(
        holding
        for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2)
        if (holding := _holding(row, mapping, row_number)) is not None
    )
    if not holdings:
        raise WorkbookBuildError("No holding rows were found below the source headers.")
    return holdings


def load_holdings(source: Path) -> tuple[Holding, ...]:
    """Load a supported plain source without carrying pre-header account text forward."""

    if not source.is_file():
        raise WorkbookBuildError("The source data file is missing.")
    extension = source.suffix.lower()
    if extension in {".csv", ".tsv"}:
        delimiter = "\t" if extension == ".tsv" else ","
        with source.open("r", encoding="utf-8-sig", newline="") as stream:
            return _from_rows([list(row) for row in csv.reader(stream, delimiter=delimiter)])
    if extension not in {".xlsx", ".xlsm"}:
        raise WorkbookBuildError("Use an .xlsx, .xlsm, .csv, or .tsv source file.")
    workbook = load_workbook(source, read_only=True, data_only=True, keep_links=False)
    try:
        visible = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
        populated = [
            sheet for sheet in visible
            if any(value is not None for row in sheet.iter_rows(values_only=True) for value in row)
        ]
        if len(populated) != 1:
            raise WorkbookBuildError("The source workbook must contain exactly one populated visible worksheet.")
        rows = [list(row) for row in populated[0].iter_rows(values_only=True)]
        return _from_rows(rows)
    finally:
        workbook.close()


def _spaced(text: str, *, pdf: bool = False) -> str:
    separator = "\u00a0" if pdf else " "
    return separator.join(text.upper())


def build_holdings_workbook(holdings: tuple[Holding, ...], config: HoldingsWorkbookConfig,
                            output_path: Path) -> Path:
    if not holdings:
        raise WorkbookBuildError("At least one holding is required.")
    if not config.report_label.strip() or not config.firm_label.strip():
        raise WorkbookBuildError("Firm and report labels are required.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = config.sheet_name[:31]
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A6"
    sheet.merge_cells("A1:C1")
    sheet.merge_cells("D1:E1")
    sheet.merge_cells("A3:E3")

    sheet["A1"] = config.firm_label
    sheet["A1"].font = Font(name="Georgia", size=17, bold=True, color=NAVY)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet["D1"] = _spaced(config.report_label)
    sheet["D1"].font = Font(name="Aptos", size=9, bold=True, color=GRAY)
    sheet["D1"].alignment = Alignment(horizontal="right", vertical="center")
    sheet.row_dimensions[1].height = 30

    navy_rule = Side(style="thin", color=NAVY)
    for cell in sheet[2]:
        cell.border = Border(bottom=navy_rule)
    sheet.row_dimensions[2].height = 5
    sheet["A3"] = f"Source: {config.source_label}"
    sheet["A3"].font = Font(name="Aptos", size=8, italic=True, color=GRAY)
    sheet["A3"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[3].height = 18
    sheet.row_dimensions[4].height = 7

    headers = ("Company", "Quantity", "Price", "Value", "% of Assets")
    for column, label in enumerate(headers, start=1):
        cell = sheet.cell(5, column, _spaced(label))
        cell.font = Font(name="Aptos", size=8, bold=True, color=GRAY)
        cell.alignment = Alignment(horizontal="right" if column > 1 else "left", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color="BFC6D0"))
    sheet.row_dimensions[5].height = 23

    ticker_font = InlineFont(rFont="Aptos", sz=10, b=True, color=NAVY)
    company_font = InlineFont(rFont="Aptos", sz=10, color=GRAY)
    first_row = 6
    for offset, holding in enumerate(holdings):
        row = first_row + offset
        company = sheet.cell(row, 1)
        company.value = CellRichText(
            TextBlock(ticker_font, holding.symbol),
            TextBlock(company_font, f"  {holding.company}"),
        )
        company.alignment = Alignment(vertical="center", wrap_text=True)
        for column, value in enumerate(
            (holding.quantity, holding.price, holding.value, holding.percent_assets), start=2
        ):
            cell = sheet.cell(row, column, value)
            cell.font = Font(name="Aptos", size=10, color=INK)
            cell.alignment = Alignment(horizontal="right", vertical="center")
        sheet.cell(row, 2).number_format = "#,##0.00"
        sheet.cell(row, 3).number_format = '$#,##0.00;[Red]($#,##0.00);-'
        sheet.cell(row, 4).number_format = '$#,##0.00;[Red]($#,##0.00);-'
        sheet.cell(row, 5).number_format = "0.00%;[Red](0.00%);-"
        if offset % 2:
            for cell in sheet[row]:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
        wraps = max(1, (len(holding.symbol) + 2 + len(holding.company) + 47) // 48)
        sheet.row_dimensions[row].height = 22 + (wraps - 1) * 12

    total_row = first_row + len(holdings)
    sheet.cell(total_row, 1, "Total")
    sheet.cell(total_row, 4, f"=SUM(D{first_row}:D{total_row - 1})")
    sheet.cell(total_row, 5, f"=SUM(E{first_row}:E{total_row - 1})")
    for column in range(1, 6):
        cell = sheet.cell(total_row, column)
        cell.font = Font(name="Aptos", size=10, bold=True, color=NAVY)
        cell.border = Border(top=navy_rule)
        cell.alignment = Alignment(horizontal="right" if column > 1 else "left", vertical="center")
    sheet.cell(total_row, 4).number_format = '$#,##0.00;[Red]($#,##0.00);-'
    sheet.cell(total_row, 5).number_format = "0.00%;[Red](0.00%);-"
    sheet.row_dimensions[total_row].height = 25

    widths = (52, 15, 15, 17, 16)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.auto_filter.ref = f"A5:E{total_row - 1}"
    sheet.print_area = f"A1:E{total_row}"
    sheet.print_title_rows = "5:5"
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = .5
    sheet.page_margins.right = .5
    sheet.page_margins.top = .5
    sheet.page_margins.bottom = .5
    sheet.oddFooter.center.text = "Gottfried & Somberg Wealth Management"
    sheet.oddFooter.right.text = "Page &P of &N"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    qa = OutputInspector().inspect(output_path)
    if not qa.approved:
        output_path.unlink(missing_ok=True)
        raise WorkbookBuildError("The generated workbook failed integrity checks.")
    return output_path


def build_holdings_snapshot(holdings: tuple[Holding, ...], config: HoldingsWorkbookConfig,
                            output_path: Path) -> Path:
    """Render the same holdings data as a compact PDF for in-app review."""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(output_path), pagesize=letter, leftMargin=.5 * inch, rightMargin=.5 * inch,
        topMargin=.45 * inch, bottomMargin=.45 * inch,
        title=config.report_label, author=config.firm_label,
    )
    firm_style = ParagraphStyle("Firm", fontName="Times-Bold", fontSize=17, leading=21, textColor=_pdf_color(NAVY))
    report_style = ParagraphStyle("Report", fontName="Helvetica-Bold", fontSize=7.5, leading=10, textColor=_pdf_color(GRAY), alignment=TA_RIGHT)
    source_style = ParagraphStyle("Source", fontName="Helvetica-Oblique", fontSize=7.5, leading=10, textColor=_pdf_color(GRAY))
    header_style = ParagraphStyle("Header", fontName="Helvetica-Bold", fontSize=6.8, leading=8, textColor=_pdf_color(GRAY))
    company_style = ParagraphStyle("Company", fontName="Helvetica", fontSize=8.4, leading=10, textColor=_pdf_color(GRAY))
    number_style = ParagraphStyle("Number", fontName="Helvetica", fontSize=8.4, leading=10, textColor=_pdf_color(INK), alignment=TA_RIGHT)
    total_style = ParagraphStyle("Total", fontName="Helvetica-Bold", fontSize=8.5, leading=10, textColor=_pdf_color(NAVY), alignment=TA_RIGHT)

    story = [
        Table(
            [[Paragraph(escape(config.firm_label), firm_style), Paragraph(_spaced(escape(config.report_label), pdf=True), report_style)]],
            colWidths=[5.1 * inch, 2.4 * inch],
            style=TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
                ("LINEBELOW", (0, 0), (-1, -1), .8, _pdf_color(NAVY)),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]),
        ),
        Spacer(1, 7),
        Paragraph(f"Source: {escape(config.source_label)}", source_style),
        Spacer(1, 11),
    ]
    data = [[
        Paragraph(_spaced("Company", pdf=True), header_style),
        Paragraph(_spaced("Quantity", pdf=True), header_style),
        Paragraph(_spaced("Price", pdf=True), header_style),
        Paragraph(_spaced("Value", pdf=True), header_style),
        Paragraph(_spaced("% of Assets", pdf=True), header_style),
    ]]
    for holding in holdings:
        company = f'<font color="#{NAVY}"><b>{escape(holding.symbol)}</b></font>&nbsp;&nbsp;{escape(holding.company)}'
        data.append([
            Paragraph(company, company_style),
            Paragraph(f"{holding.quantity:,.2f}", number_style),
            Paragraph(f"${holding.price:,.2f}", number_style),
            Paragraph(f"${holding.value:,.2f}", number_style),
            Paragraph(f"{holding.percent_assets:.2%}", number_style),
        ])
    data.append([
        Paragraph("<b>Total</b>", company_style), "", "",
        Paragraph(f"${sum(item.value for item in holdings):,.2f}", total_style),
        Paragraph(f"{sum(item.percent_assets for item in holdings):.2%}", total_style),
    ])
    table = Table(data, colWidths=[3.5 * inch, 1.0 * inch, .9 * inch, 1.15 * inch, .95 * inch], repeatRows=1)
    styles = [
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, 0), .45, _pdf_color("BFC6D0")),
        ("LINEABOVE", (0, -1), (-1, -1), .8, _pdf_color(NAVY)),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
    ]
    for row in range(2, len(data) - 1, 2):
        styles.append(("BACKGROUND", (0, row), (-1, row), _pdf_color(LIGHT)))
    table.setStyle(TableStyle(styles))
    story.append(table)
    doc.build(story)
    qa = OutputInspector().inspect(output_path)
    if not qa.approved:
        output_path.unlink(missing_ok=True)
        raise WorkbookBuildError("The workbook preview failed integrity checks.")
    return output_path
