"""Structural input validation performed after the global privacy gate."""

from __future__ import annotations

import csv
import json
import zipfile
from dataclasses import dataclass
from enum import StrEnum
from pathlib import Path

from core.workflows import ReportWorkflow


class ValidationSeverity(StrEnum):
    ERROR = "error"
    WARNING = "warning"


@dataclass(frozen=True, slots=True)
class InputIssue:
    severity: ValidationSeverity
    slot_id: str
    filename: str
    code: str
    message: str


@dataclass(frozen=True, slots=True)
class InputValidationResult:
    approved: bool
    issues: tuple[InputIssue, ...] = ()

    @property
    def errors(self) -> tuple[InputIssue, ...]:
        return tuple(issue for issue in self.issues if issue.severity is ValidationSeverity.ERROR)

    @property
    def warnings(self) -> tuple[InputIssue, ...]:
        return tuple(issue for issue in self.issues if issue.severity is ValidationSeverity.WARNING)


class InputValidator:
    """Validate declared upload contracts and file structure without AI."""

    def __init__(self, max_file_size_bytes: int = 100 * 1024 * 1024):
        self.max_file_size_bytes = max_file_size_bytes

    def validate(
        self,
        workflow: ReportWorkflow,
        selections: dict[str, tuple[Path, ...]],
    ) -> InputValidationResult:
        issues: list[InputIssue] = []
        requirements = {
            item.requirement_id: item
            for item in workflow.required_uploads + workflow.optional_uploads
        }
        for required in workflow.required_uploads:
            if not selections.get(required.requirement_id):
                issues.append(
                    InputIssue(
                        ValidationSeverity.ERROR,
                        required.requirement_id,
                        "",
                        "required_upload_missing",
                        f"{required.label} is required.",
                    )
                )

        for slot_id, paths in selections.items():
            if slot_id == "custom":
                accepted_extensions = None
            elif slot_id not in requirements:
                issues.append(
                    InputIssue(
                        ValidationSeverity.ERROR,
                        slot_id,
                        "",
                        "unknown_upload_slot",
                        "The upload does not belong to this report workflow.",
                    )
                )
                continue
            else:
                accepted_extensions = requirements[slot_id].accepted_extensions

            for path in paths:
                issues.extend(self._validate_file(slot_id, path, accepted_extensions))

        return InputValidationResult(
            approved=not any(issue.severity is ValidationSeverity.ERROR for issue in issues),
            issues=tuple(issues),
        )

    def _validate_file(
        self,
        slot_id: str,
        path: Path,
        accepted_extensions: tuple[str, ...] | None,
    ) -> list[InputIssue]:
        issues: list[InputIssue] = []

        def error(code: str, message: str) -> None:
            issues.append(InputIssue(ValidationSeverity.ERROR, slot_id, path.name, code, message))

        if not path.is_file():
            error("file_missing", "The selected file no longer exists.")
            return issues
        try:
            size = path.stat().st_size
        except OSError:
            error("file_unreadable", "The selected file cannot be read.")
            return issues
        if size == 0:
            error("file_empty", "The selected file is empty.")
            return issues
        if size > self.max_file_size_bytes:
            error("file_too_large", "The selected file exceeds the 100 MB limit.")
            return issues

        extension = path.suffix.lower()
        if accepted_extensions is not None and extension not in accepted_extensions:
            allowed = ", ".join(accepted_extensions)
            error("file_type_not_allowed", f"Use one of these file types: {allowed}.")
            return issues

        try:
            self._validate_structure(path, extension)
        except (OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
            error("file_structure_invalid", str(exc))
        return issues

    def _validate_structure(self, path: Path, extension: str) -> None:
        if extension == ".json":
            with path.open("r", encoding="utf-8-sig") as stream:
                payload = json.load(stream)
            if not isinstance(payload, (dict, list)):
                raise ValueError("The JSON file must contain an object or list.")
            return
        if extension in {".csv", ".tsv"}:
            delimiter = "\t" if extension == ".tsv" else ","
            with path.open("r", encoding="utf-8-sig", errors="strict", newline="") as stream:
                rows = csv.reader(stream, delimiter=delimiter)
                if next(rows, None) is None:
                    raise ValueError("The data file contains no rows.")
            return
        if extension in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
            try:
                if not workbook.sheetnames:
                    raise ValueError("The workbook contains no worksheets.")
                if not any(
                    value is not None
                    for sheet in workbook.worksheets
                    for row in sheet.iter_rows(values_only=True)
                    for value in row
                ):
                    raise ValueError("The workbook contains no data.")
            finally:
                workbook.close()
            return
        if extension == ".pdf":
            from pypdf import PdfReader

            reader = PdfReader(path)
            if reader.is_encrypted:
                raise ValueError("Encrypted PDFs are not supported.")
            if not reader.pages:
                raise ValueError("The PDF contains no pages.")
            return
        if extension == ".pptx":
            from pptx import Presentation

            if not Presentation(path).slides:
                raise ValueError("The presentation contains no slides.")
            return
        if extension == ".docx":
            with zipfile.ZipFile(path) as archive:
                if "word/document.xml" not in archive.namelist():
                    raise ValueError("The Word document structure is invalid.")
            return
        if extension in {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}:
            from PIL import Image

            with Image.open(path) as image:
                image.verify()
            return
        raise ValueError(f"The {extension or 'unknown'} file type is not structurally supported.")
